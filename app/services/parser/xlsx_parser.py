"""XLSX 바이너리에서 텍스트를 추출한다.

시트별로 셀 텍스트를 순서대로 이어붙여 하나의 문자열로 반환한다. 서식/차트/
이미지 보존은 다루지 않는다(YAGNI, 후속 단계).
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.core.errors import ParserError


def extract_text(data: bytes) -> str:
    """XLSX 바이트에서 시트별 셀 텍스트를 순서대로 이어붙여 반환한다.

    `read_only=True` 로 열어 대용량 워크북도 전체를 메모리에 올리지 않고
    스트리밍으로 순회한다.

    Raises:
        ParserError: 손상된 파일이거나 XLSX 형식이 아닌 경우.
    """
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        try:
            lines: list[str] = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    cells = [str(cell.value) for cell in row if cell.value is not None]
                    if cells:
                        lines.append(" ".join(cells))
        finally:
            workbook.close()
    except Exception as exc:
        raise ParserError(f"failed to parse XLSX: {exc}") from exc
    return "\n".join(lines)
